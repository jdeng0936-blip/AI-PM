"""
app/routers/export.py — 数据导出 API

GET /api/v1/export/daily-reports — 按日期范围导出日报为 Excel (.xlsx)
"""
import io
from datetime import date, timedelta
from typing import Optional

from fastapi import APIRouter, Depends, Query
from fastapi.responses import StreamingResponse
from sqlalchemy import select, and_
from sqlalchemy.ext.asyncio import AsyncSession
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from app.database import get_db
from app.middleware.rbac import require_role
from app.models.daily_report import DailyReport
from app.models.user import User, UserRole

router = APIRouter(prefix="/api/v1/export", tags=["数据导出"])


@router.get("/daily-reports")
async def export_daily_reports(
    start_date: Optional[date] = Query(None, description="起始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    db: AsyncSession = Depends(get_db),
    _admin=Depends(require_role(UserRole.admin, UserRole.manager)),
):
    """
    导出日报为 Excel (.xlsx)，列对齐原始 Excel 日报表。
    默认导出最近 7 天。
    """
    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=6)

    stmt = (
        select(DailyReport, User.name, User.department)
        .join(User, DailyReport.user_id == User.id)
        .where(and_(
            DailyReport.report_date >= start_date,
            DailyReport.report_date <= end_date,
        ))
        .order_by(DailyReport.report_date.desc(), DailyReport.ai_score.desc())
    )
    rows = (await db.execute(stmt)).all()

    # ── 生成 Excel ──
    wb = Workbook()
    ws = wb.active
    ws.title = "AI日报汇总"

    # 表头样式
    header_font = Font(name="微软雅黑", bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="3B82F6", end_color="3B82F6", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin", color="D1D5DB"),
        right=Side(style="thin", color="D1D5DB"),
        top=Side(style="thin", color="D1D5DB"),
        bottom=Side(style="thin", color="D1D5DB"),
    )

    headers = [
        "日期", "姓名", "部门", "AI评分", "状态",
        "今日任务", "完成进度", "验收标准", "所需支持",
        "核心卡点", "解决方案", "预计解决", "验收人",
        "Git版本", "AI评语",
    ]

    # 列宽
    widths = [12, 8, 10, 8, 6, 35, 8, 20, 15, 20, 20, 12, 8, 12, 30]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[ws.cell(1, i).column_letter].width = w

    # 写入表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 写入数据
    pass_fill = PatternFill(start_color="DCFCE7", fill_type="solid")
    fail_fill = PatternFill(start_color="FEE2E2", fill_type="solid")
    data_align = Alignment(vertical="center", wrap_text=True)

    for row_idx, r in enumerate(rows, 2):
        pc = r.DailyReport.parsed_content or {}
        is_pass = r.DailyReport.pass_check

        values = [
            str(r.DailyReport.report_date),
            r.name,
            r.department,
            r.DailyReport.ai_score,
            "合格" if is_pass else "退回",
            pc.get("tasks", ""),
            f"{pc.get('progress', '')}%"if pc.get("progress") is not None else "",
            pc.get("acceptance_criteria", ""),
            pc.get("support_needed", ""),
            pc.get("blocker", ""),
            pc.get("next_step", ""),
            pc.get("eta", ""),
            pc.get("reviewer", ""),
            pc.get("git_version", ""),
            r.DailyReport.ai_comment or "",
        ]

        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.alignment = data_align
            cell.border = thin_border
            if is_pass is not None:
                cell.fill = pass_fill if is_pass else fail_fill

    # 冻结首行
    ws.freeze_panes = "A2"

    # 输出
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"AI日报_{start_date}_{end_date}.xlsx"
    # RFC 5987: use filename* for non-ASCII chars
    from urllib.parse import quote
    encoded_filename = quote(filename)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )


@router.get("/daily-reports-csv")
async def export_daily_reports_csv(
    start_date: Optional[date] = Query(None, description="起始日期"),
    end_date: Optional[date] = Query(None, description="结束日期"),
    db: AsyncSession = Depends(get_db),
    _admin=Depends(require_role(UserRole.admin, UserRole.manager)),
):
    """
    导出日报为 CSV 格式。
    """
    import csv

    if not end_date:
        end_date = date.today()
    if not start_date:
        start_date = end_date - timedelta(days=6)

    stmt = (
        select(DailyReport, User.name, User.department)
        .join(User, DailyReport.user_id == User.id)
        .where(and_(
            DailyReport.report_date >= start_date,
            DailyReport.report_date <= end_date,
        ))
        .order_by(DailyReport.report_date.desc(), DailyReport.ai_score.desc())
    )
    rows = (await db.execute(stmt)).all()

    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([
        "日期", "姓名", "部门", "AI评分", "状态",
        "今日任务", "完成进度", "验收标准", "核心卡点",
        "解决方案", "AI评语",
    ])

    for r in rows:
        pc = r.DailyReport.parsed_content or {}
        writer.writerow([
            str(r.DailyReport.report_date),
            r.name,
            r.department,
            r.DailyReport.ai_score,
            "合格" if r.DailyReport.pass_check else "退回",
            pc.get("tasks", ""),
            f"{pc.get('progress', '')}%",
            pc.get("acceptance_criteria", ""),
            pc.get("blocker", ""),
            pc.get("next_step", ""),
            r.DailyReport.ai_comment or "",
        ])

    csv_bytes = output.getvalue().encode("utf-8-sig")  # BOM for Excel compatibility
    from urllib.parse import quote
    filename = f"AI日报_{start_date}_{end_date}.csv"
    encoded_filename = quote(filename)

    return StreamingResponse(
        io.BytesIO(csv_bytes),
        media_type="text/csv; charset=utf-8",
        headers={
            "Content-Disposition": f"attachment; filename*=UTF-8''{encoded_filename}",
        },
    )

